import openpyxl
from DBPools import OPMysql


MYSQLINFO = {
    "host": 'localhost',
    "user": 'root',
    "passwd": '123456',
    "db": 'pro_info',
    "port": 3306,
    "charset": 'utf8mb4'
}
conn = OPMysql(MYSQLINFO)

sql = 'SELECT * FROM products WHERE have_picture=1'
results = conn.op_select_all(sql)
if results:
    workbook = openpyxl.Workbook()
    # wb=openpyxl.Workbook(encoding='UTF-8')
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    # worksheet2 = workbook.create_sheet()  # 默认插在工作簿末尾
    # worksheet2 = workbook.create_sheet(0)  #插入在工作簿的第一个位置
    # worksheet2.title = "Sheet2"
    field = ['产品名称', 'Model', '分类', '现价', '原价', '尺码', '颜色', '图片', '描述']

    for i in range(len(field)):
        worksheet.cell(1, i + 1, field[i])

    row_num = 2
    for row in results:
        field_value_lst = []
        pro_title = row['pro_title']
        field_value_lst.append(pro_title)
        pro_model = row['pro_model']
        field_value_lst.append(pro_model)
        category = row['category']
        if category == 'Cocktail Dresses':
            category = 'Cocktail'
        elif category == 'Homecoming Dresses':
            category = 'Homecoming'
        elif category == 'Prom Dresses':
            category = 'Prom'
        elif category == 'Homecoming Dresses':
            category = 'Homecoming'
        field_value_lst.append(category) 
        old_price = row['pro_price']
        new_price = int(old_price) - 5
        field_value_lst.append(new_price)
        field_value_lst.append(old_price)
        pro_size = ''
        field_value_lst.append(pro_size)
        pro_color = ''
        field_value_lst.append(pro_color)
        pro_image = 'media/' + pro_model + '.jpg'
        field_value_lst.append(pro_image) 
        pro_desc = row['pro_desc']
        field_value_lst.append(pro_desc)

        for i in range(len(field_value_lst)):
            worksheet.cell(row_num, i + 1, field_value_lst[i])
        row_num += 1

workbook.save(filename='product_data.xlsx')
